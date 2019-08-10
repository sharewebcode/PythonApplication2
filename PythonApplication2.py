import pandas as pd
import openpyxl
from openpyxl.chart import Reference, Series, BarChart
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side, Color, PatternFill

#df = pd.read_excel('pandasSample.xlsx')

# From : To : Step - 짝수행 추출
#df[::2].to_excel('even.xlsx')

# 홀수행 추출
#df[1::2].to_excel('odd.xlsx')

df_Sam = pd.read_excel('2017Sam.xlsx')
df_Sam.set_index('date', inplace=True)

df_LG = pd.read_excel('2017LG.xlsx')
df_LG.set_index('date', inplace=True)

df_merge = pd.DataFrame()

df_merge['삼성'] = df_Sam['total']
df_merge['LG'] = df_LG['total']

df_merge.to_excel('merge.xlsx')

wb = openpyxl.load_workbook('merge.xlsx')
sheet = wb.active

#sum_Sam = sum([row[0].value for row in sheet['B2':'B13']])
#sheet['B14'].value=sum_Sam

#sum_LG = sum([row[0].value for row in sheet['C2':'C13']])
#sheet['C14'].value=sum_LG

for row in sheet['A2:A13']:
    for cell in row:
        cell.number_format = 'yyyy-mm'

sheet['A14'].value = '합계'
sheet['B14'].value = '=SUM(B2:B13)'
sheet['C14'].value = '=SUM(C2:C13)'

font_11 = Font(name='맑은 고딕', size=11, bold=True)
font_15 = Font(name='맑은 고딕', size=15, bold=True)

align_center = Alignment(horizontal='center', vertical='center')
align_vcenter = Alignment(vertical='center')

border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

fill_orange = PatternFill(patternType='solid', fgColor=Color('FFC000'))
fill_lightgrey = PatternFill(patternType='solid', fgColor=('D3D3D3'))

cell_sum = sheet['A14']
cell_sum.font = font_15
cell_sum.alignment = align_center
cell_sum.border = border_thin
cell_sum.fill = fill_orange

for row in sheet['B2:C14']:
    for cell in row:        
        cell.border = border_thin
        cell.number_format = '0.00'

for row in sheet['B14:C14']:
    for cell in row:
        cell.alignment = align_vcenter
        cell.fill = fill_orange


chart = BarChart()
chart.title = '2017년 월별 광고비 (억원)'

values = Reference(sheet, range_string='Sheet1!B2:B13')
series = Series(values, title='SamSung')
chart.append(series)

values = Reference(sheet, range_string='Sheet1!C2:C13')
series = Series(values, title='LG전자')
chart.append(series)

sheet.add_chart(chart, 'E1')

wb.save('merge.xlsx')


